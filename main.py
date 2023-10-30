import openpyxl

file = openpyxl.load_workbook("inventory.xlsx")
list = file["Sheet1"]

supplier = {}
total_value = {}
under10inv = {}

for row in range(2,list.max_row+1) :
    supplier_name = list.cell(row,4).value
    inventory = list.cell(row,2).value
    price = list.cell(row,3).value
    product_no = list.cell(row,1).value
    new_list = list.cell(row,5)
    
    # print(f"{supplier_name}:{inventory}")
    #* execercise 1
    if supplier_name in supplier :
        supplier[supplier_name] = supplier[supplier_name]+1
    else :
        print("adding supplier")
        supplier[supplier_name] = 1
        
    #* execercise 2
    if supplier_name in total_value :
        current = total_value[supplier_name]
        total_value[supplier_name] = current + price * inventory
    else :
        total_value[supplier_name] = price * inventory

    #* execercise 3
    if inventory  < 10 :
        under10inv[product_no] = inventory

    #* execercise 4
    curr_val = inventory * price
    new_list.value = curr_val

file.save("inventory_new.xlsx")

print(supplier)
print(total_value)
print(under10inv)